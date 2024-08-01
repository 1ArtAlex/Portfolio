import telebot
from telebot import types
from openpyxl import load_workbook
import sqlite3
import traceback
import time
from apscheduler.schedulers.background import BackgroundScheduler
import re


TOKEN = 'BOT_TOKEN'

bot = telebot.TeleBot(TOKEN)

cities_data = {}
file_uploaded = False
admin_broadcast = False
ADMIN_ID = ['ADMIN_FIRST_ID', 'ADMIN_SECOND_ID']
admin_chat_id = -ADMIN_CHAT_ID

# Создание базы данных
conn = sqlite3.connect('bot_database.db')
cursor = conn.cursor()

cursor.execute('''CREATE TABLE IF NOT EXISTS users
                  (user_id INTEGER PRIMARY KEY, 
                   name TEXT, 
                   tg_link TEXT, 
                   status INTEGER DEFAULT 1)''')

cursor.execute('''CREATE TABLE IF NOT EXISTS user_city_history_pickup
                  (user_id INTEGER, 
                   city TEXT, 
                   price_A REAL, 
                   price_B REAL, 
                   FOREIGN KEY(user_id) REFERENCES users(user_id))''')

cursor.execute('''CREATE TABLE IF NOT EXISTS user_city_history_delivery
                  (user_id INTEGER, 
                  city TEXT, 
                  price_A REAL, 
                  price_B REAL, 
                  price_C REAL, 
                  FOREIGN KEY(user_id) REFERENCES users(user_id))''')

cursor.execute('''CREATE TABLE IF NOT EXISTS new_city_pickup
                  (city TEXT PRIMARY KEY, 
                  price_A REAL, 
                  price_B REAL)''')

cursor.execute('''CREATE TABLE IF NOT EXISTS new_city_delivery
                  (city TEXT PRIMARY KEY, 
                  price_A REAL, 
                  price_B REAL, 
                  price_C REAL)''')


# Добавление пользователей в базу данных
def add_user_to_database(user_id, name, tg_link):
    conn = sqlite3.connect('bot_database.db')
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM users WHERE user_id = ?", (user_id,))
    result = cursor.fetchone()
    if result[0] == 0:
        cursor.execute("INSERT INTO users (user_id, name, tg_link) VALUES (?, ?, ?)", (user_id, name, tg_link))
        conn.commit()

    conn.close()


# Получение цен по городам из разных листов в excel таблице
def get_prices_by_shipment_type(city, shipment_type):
    file_path = 'бот.xlsx'

    if shipment_type == 'доставка':
        sheet_name = 'доставка'
        price_columns = [1, 3, 5]
    elif shipment_type == 'самовывоз':
        sheet_name = 'самовывоз'
        price_columns = [1, 3]
    else:
        return None

    wb = load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]

    prices = None
    for row in sheet.iter_rows(values_only=True):
        if row[0] == city:
            prices = [row[col] for col in price_columns]
            prices = [0 if price is None else price for price in prices]
            break

    return prices


# Добавление посмотренных городов пользователями в базу данных
def add_city_to_history(user_id, city, shipment_type):
    conn = sqlite3.connect('bot_database.db')
    cursor = conn.cursor()

    if shipment_type == 'самовывоз':
        table_name = 'user_city_history_pickup'
    elif shipment_type == 'доставка':
        table_name = 'user_city_history_delivery'
    else:
        return

    cursor.execute(f"SELECT COUNT(*) FROM {table_name} WHERE user_id = ? AND city = ?", (user_id, city))
    existing_count = cursor.fetchone()[0]

    if existing_count > 0:
        cursor.execute(f"DELETE FROM {table_name} WHERE user_id = ? AND city = ?", (user_id, city))

    cursor.execute(f"SELECT COUNT(*) FROM {table_name} WHERE user_id = ?", (user_id,))
    count = cursor.fetchone()[0]

    if count >= 3:
        cursor.execute(f"SELECT city FROM {table_name} WHERE user_id = ?", (user_id,))
        old_city = cursor.fetchone()[0]
        cursor.execute(f"DELETE FROM {table_name} WHERE user_id = ? AND city = ?", (user_id, old_city))

    prices = get_prices_by_shipment_type(city, shipment_type)
    if prices is not None:
        if shipment_type == 'доставка':
            cursor.execute(
                f"INSERT INTO {table_name} (user_id, city, price_A, price_B, price_C) "
                "VALUES (?, ?, ?, ?, ?)", (user_id, city, *prices))
        else:
            cursor.execute(
                f"INSERT INTO {table_name} (user_id, city, price_A, price_B) "
                "VALUES (?, ?, ?, ?)", (user_id, city, *prices))
    else:
        cursor.execute(f"INSERT INTO {table_name} (user_id, city) VALUES (?, ?)", (user_id, city))

    conn.commit()
    conn.close()


# Команда контакты
@bot.message_handler(commands=['contacts'])
def contacts_handler(message):
    if is_user_blocked(message.from_user.id):
        bot.send_message(message.chat.id, "Для доступа к данному ресурсу обратитесь к ответственному менеджеру. "
                                          "\n[Связаться с менеджером](https://t.me/RybinVladimirNN)"
                         , parse_mode='Markdown')
    else:
        user_id = message.chat.id
        send_cities(user_id, message.message_id, 'контакты')


# Команда доставка
@bot.message_handler(commands=['delivery'])
def delivery_start(message):
    if is_user_blocked(message.from_user.id):
        bot.send_message(message.chat.id, "Для доступа к данному ресурсу обратитесь к ответственному менеджеру. "
                                          "\n[Связаться с менеджером](https://t.me/RybinVladimirNN)"
                         , parse_mode='Markdown')
    else:
        send_cities(message.chat.id, message.message_id, 'доставка')


# Команда самовывоз
@bot.message_handler(commands=['pickup'])
def delivery_start(message):
    if is_user_blocked(message.from_user.id):
        bot.send_message(message.chat.id, "Для доступа к данному ресурсу обратитесь к ответственному менеджеру. "
                                          "\n[Связаться с менеджером](https://t.me/RybinVladimirNN)"
                         , parse_mode='Markdown')
    else:
        send_cities(message.chat.id, message.message_id, 'самовывоз')


# Команда старт
@bot.message_handler(commands=['start'])
def start(message):
    user_id = message.chat.id
    first_name = message.from_user.first_name
    tg_username = message.from_user.username
    tg_link = f"https://t.me/{tg_username}"
    add_user_to_database(user_id, first_name, tg_link)

    admin_message = f"Пользователь {first_name} (<a href='tg://user?id={user_id}'>" \
                    f"https://t.me/{message.from_user.username}</a>) зашел в бот ({user_id})"

    bot.send_message(admin_chat_id, admin_message, parse_mode='HTML')

    if is_user_blocked(user_id):
        bot.send_message(user_id, "Для доступа к данному ресурсу обратитесь к ответственному менеджеру. "
                                  "\n[Связаться с менеджером](https://t.me/RybinVladimirNN)", parse_mode='Markdown')
    else:
        markup = types.InlineKeyboardMarkup(row_width=1)
        btn1 = types.InlineKeyboardButton('Самовывоз', callback_data='самовывоз')
        btn2 = types.InlineKeyboardButton('Доставка', callback_data='доставка')
        btn3 = types.InlineKeyboardButton('Контакты', callback_data='контакты')
        markup.add(btn1, btn2, btn3)

        bot.send_message(message.chat.id,
                         "<b>Компания Продлогистика занимается оптовой поставкой сахарного песка в мешках для "
                         "производителей и фасованного по 1 и 5кг собственного производства для торговых сетей</b>"
                         "\n\n✅ Предлагаем отгрузки сахара в мешках со всех сахарных заводов: "
                         "\n• Краснодарского края "
                         "\n• Центральной части   "
                         "\n• Поволжья  "
                         "\n• Татарстана  "
                         "\n• Башкирии  "
                         "\n• Алтайского края "
                         "\n\nИ своих складов с городах:  "
                         "\n• Казань"
                         "\n• Свердловская область г. Богданович "
                         "\n• Йошкар-Ола"
                         "\n• Ижевск"
                         "\n• Челябинск"
                         "\n• Барнаул"
                         "\n• Электросталь Московской области"
                         "\n• Санкт-Петербург"
                         "\n\n✅ Поставки осуществляются автомобильным и железнодорожным транспортом "
                         "(вагоны и контейнеры) в любой регион страны и ближнего "
                         "зарубежья, используя собственную службу доставки.  "
                         "\n\n✅ Данный бот позволит оперативно получить информацию об актуальных ценах компании. "
                         "\n\n✅ Выберете интересующий Вас раздел:",
                         reply_markup=markup, parse_mode='HTML')


# Команда меню
@bot.message_handler(commands=['menu'])
def main_menu1(message):
    if is_user_blocked(message.from_user.id):
        bot.send_message(message.chat.id, "Для доступа к данному ресурсу обратитесь к ответственному менеджеру. "
                                          "\n[Связаться с менеджером](https://t.me/RybinVladimirNN)"
                         , parse_mode='Markdown')
    else:
        markup = types.InlineKeyboardMarkup(row_width=1)
        btn1 = types.InlineKeyboardButton('Самовывоз', callback_data='самовывоз')
        btn2 = types.InlineKeyboardButton('Доставка', callback_data='доставка')
        btn3 = types.InlineKeyboardButton('Контакты', callback_data='контакты')
        markup.add(btn1, btn2, btn3)

        bot.send_message(message.chat.id,
                         "Выберете интересующий Вас раздел",
                         reply_markup=markup)


# Функция меню
def main_menu(message):
    markup = types.InlineKeyboardMarkup(row_width=1)
    btn1 = types.InlineKeyboardButton('Самовывоз', callback_data='самовывоз')
    btn2 = types.InlineKeyboardButton('Доставка', callback_data='доставка')
    btn3 = types.InlineKeyboardButton('Контакты', callback_data='контакты')
    markup.add(btn1, btn2, btn3)

    bot.send_message(message.chat.id,
                     "Выберете интересующий Вас раздел",
                     reply_markup=markup)


# Команда админ
@bot.message_handler(commands=['admin'])
def handle_admin_command(message):
    if str(message.from_user.id) in ADMIN_ID:
        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
        markup.add('Загрузить файл', 'Рассылка', 'Клиенты', 'Отмена')
        bot.send_message(message.chat.id, "Нажмите кнопку ниже, чтобы выбрать действие.", reply_markup=markup)
    else:
        bot.send_message(message.chat.id, "У вас нет прав для использования этой команды.")


# Обработка кнопок из админ-панели
@bot.message_handler(content_types=['text'])
def handle_text(message):
    global admin_broadcast

    if message.text == 'Загрузить файл':
        bot.send_message(message.chat.id, "Пожалуйста, отправьте файл.")
    elif message.text == 'Отмена':
        admin_broadcast = False
        remove_markup = types.ReplyKeyboardRemove()
        bot.send_message(message.chat.id, "Вы вышли из админ-панели", reply_markup=remove_markup)
    elif message.text == 'Рассылка':
        admin_broadcast = True
        bot.send_message(message.chat.id, "Введите сообщение для рассылки или отправьте фото. "
                                          "Для отмены нажмите 'Отмена'.")
    elif message.text == 'Клиенты':
        send_clients_list(message)
    elif admin_broadcast:
        broadcast_message(message)
    else:
        forward_message_to_admin(message)


# Отправка списка кнопок с клиентами
def send_clients_list(message):
    conn = sqlite3.connect('bot_database.db')
    cursor = conn.cursor()

    cursor.execute('SELECT user_id, name, tg_link, status FROM users')
    clients = cursor.fetchall()
    conn.close()

    if clients:
        markup = types.InlineKeyboardMarkup()

        for user_id, name, tg_link, status in clients:
            status_text = "✅" if status == 1 else "❌"
            callback_data = f"user_{user_id}"
            button_text = f"{status_text} {name} (ID: {user_id}):"

            markup.row(types.InlineKeyboardButton(button_text, callback_data=callback_data))

        bot.send_message(message.chat.id, "Ваши клиенты:"
                                          "\n✅ - пользователь не заблокирован (нажмите, чтобы заблокировать)"
                                          "\n❌ - пользователь заблокирован (нажмите, чтобы разблокировать)"
                         , reply_markup=markup, parse_mode="HTML")
    else:
        bot.send_message(message.chat.id, "Клиенты не найдены.")


# Сообщения о блокировке и разблокировке клиентов
@bot.callback_query_handler(func=lambda call: call.data.startswith('user_'))
def handle_user_callback(call):
    user_id = call.data.split('_')[1]
    current_status = get_user_status(user_id)

    if current_status == 1:
        block_user(user_id)
        remove_markup = types.ReplyKeyboardRemove()
        bot.send_message(call.message.chat.id, f"Пользователь с ID {user_id} заблокирован."
                         , reply_markup=remove_markup)
    else:
        unblock_user(user_id)
        remove_markup = types.ReplyKeyboardRemove()
        bot.send_message(call.message.chat.id, f"Пользователь с ID {user_id} разблокирован."
                         , reply_markup=remove_markup)


# Получение статуса клиентов
def get_user_status(user_id):
    conn = sqlite3.connect('bot_database.db')
    cursor = conn.cursor()

    cursor.execute("SELECT status FROM users WHERE user_id = ?", (user_id,))
    status = cursor.fetchone()[0]

    conn.close()

    return status


# Блокировка клиентов
def block_user(user_id):
    conn = sqlite3.connect('bot_database.db')
    cursor = conn.cursor()

    cursor.execute("UPDATE users SET status = 0 WHERE user_id = ?", (user_id,))
    cursor.execute("DELETE FROM user_city_history_pickup WHERE user_id = ?", (user_id,))
    cursor.execute("DELETE FROM user_city_history_delivery WHERE user_id = ?", (user_id,))

    conn.commit()
    conn.close()


# Разблокировка клиентов
def unblock_user(user_id):
    conn = sqlite3.connect('bot_database.db')
    cursor = conn.cursor()

    cursor.execute("UPDATE users SET status = 1 WHERE user_id = ?", (user_id,))

    conn.commit()
    conn.close()


# Если клиент заблокирован
def is_user_blocked(user_id):
    conn = sqlite3.connect('bot_database.db')
    cursor = conn.cursor()

    cursor.execute("SELECT status FROM users WHERE user_id = ?", (user_id,))
    status = cursor.fetchone()[0]

    conn.close()

    return status == 0


# Заперт на нажатии кнопок заблокированным клиентам
@bot.callback_query_handler(func=lambda call: is_user_blocked(call.from_user.id))
def blocked_user_callback(call):
    bot.send_message(call.message.chat.id, "Для доступа к данному ресурсу обратитесь к ответственному менеджеру. "
                                           "\n[Связаться с менеджером](https://t.me/RybinVladimirNN)"
                     , parse_mode='Markdown')


# Рассылка
@bot.message_handler(content_types=['photo', 'video', 'animation'])
def handle_photo(message):
    global admin_broadcast

    if admin_broadcast:
        broadcast_message(message)
    else:
        forward_message_to_admin(message)


# Сама рассылка
def broadcast_message(message):
    global admin_broadcast

    conn = sqlite3.connect('bot_database.db')
    cursor = conn.cursor()
    cursor.execute('SELECT user_id, status FROM users')
    user_statuses = cursor.fetchall()
    conn.close()

    formatted_text = format_bold_text(message.text) if message.content_type == 'text' else None

    for user_id, status in user_statuses:
        if status == 1:
            try:
                if message.content_type == 'text':
                    bot.send_message(user_id, formatted_text, parse_mode="Markdown")
                elif message.content_type == 'photo':
                    bot.send_photo(user_id, message.photo[-1].file_id, caption=format_bold_text(message.caption)
                                   , parse_mode="Markdown")
                elif message.content_type == 'video':
                    bot.send_video(user_id, message.video.file_id, caption=format_bold_text(message.caption)
                                   , parse_mode="Markdown")
                elif message.content_type == 'animation':
                    bot.send_animation(user_id, message.document.file_id, caption=format_bold_text(message.caption)
                                       , parse_mode="Markdown")
            except Exception as e:
                print(f"Не удалось отправить сообщение пользователю {user_id}: {e}")

    remove_markup = types.ReplyKeyboardRemove()
    bot.send_message(message.chat.id, "Сообщение было отправлено всем пользователям."
                     , reply_markup=remove_markup)
    admin_broadcast = False


# Форматирование текста на жирным
def format_bold_text(text):
    if text:
        return re.sub(r'\*(.*?)\*', r'*\1*', text)
    return text


# Получение случайных сообщений от клиентов
@bot.message_handler(func=lambda message: message.text and not (message.text.startswith('/')
                                                                or message.text == 'Отмена'
                                                                or message.text == 'Загрузить файл'
                                                                or message.text == 'Рассылка')
                                          and message.chat.type == 'private', content_types=['text'])
def handle_text_message(message):
    forward_message_to_admin(message)


# Отправка полученных сообщений в чат админам
def forward_message_to_admin(message):
    user_id = message.from_user.id
    if not is_user_blocked(user_id) and str(user_id) not in ADMIN_ID:
        first_name = message.from_user.first_name
        tg_username = message.from_user.username
        tg_link = f"https://t.me/{tg_username}"
        admin_message = f"Пользователь {first_name} (<a href='tg://user?id={user_id}'>{tg_link}</a>) " \
                        f"отправил сообщение:\n\n{message.text}"
        bot.send_message(admin_chat_id, admin_message, parse_mode="HTML")


# Загрузка таблицы с данными в бота
@bot.message_handler(content_types=['document'])
def handle_document(message):
    if str(message.from_user.id) in ADMIN_ID:
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)

        filename = 'бот.xlsx'
        with open(filename, 'wb') as new_file:
            new_file.write(downloaded_file)

        global file_uploaded
        file_uploaded = True

        remove_markup = types.ReplyKeyboardRemove()
        bot.send_message(message.chat.id, "Файл успешно сохранен.", reply_markup=remove_markup)

        update_database_from_excel()
    else:
        bot.send_message(message.chat.id, "У вас нет прав для загрузки файла.")


# Обновление актуальных цен в базе данных
def update_database_from_excel():
    if file_uploaded:
        clear_tables()

        wb = load_workbook('бот.xlsx', data_only=True)

        for sheet_name in ['самовывоз', 'доставка']:
            sheet = wb[sheet_name]
            shipment_type = sheet_name.split()[0]

            start_row = 3 if shipment_type == 'самовывоз' else 2

            for row in sheet.iter_rows(min_row=start_row, values_only=True):
                city = row[0]
                prices = extract_prices(row, shipment_type)

                if city and any(prices):
                    if shipment_type == 'доставка':
                        table_name = 'new_city_delivery'
                    elif shipment_type == 'самовывоз':
                        table_name = 'new_city_pickup'

                    add_city_to_database(table_name, city, shipment_type, prices)

        compare_prices_and_update()
        compare_prices_and_update_pickup()


def clear_tables():
    conn = sqlite3.connect('bot_database.db')
    cursor = conn.cursor()

    cursor.execute("DELETE FROM new_city_delivery")
    cursor.execute("DELETE FROM new_city_pickup")

    conn.commit()
    conn.close()


# Получение индекса цен
def extract_prices(row, shipment_type):
    if shipment_type == 'доставка':
        price_columns = [1, 3, 5]
    elif shipment_type == 'самовывоз':
        price_columns = [1, 3]

    prices = [row[col] for col in price_columns]
    prices = [0 if price is None else price for price in prices]
    return prices


# Само обновление
def add_city_to_database(table_name, city, shipment_type, prices):
    conn = sqlite3.connect('bot_database.db')
    cursor = conn.cursor()

    prices = [round(price, 1) for price in prices]

    if shipment_type == 'доставка':
        cursor.execute(
            f"CREATE TABLE IF NOT EXISTS {table_name} (city TEXT PRIMARY KEY, price_A REAL, price_B REAL, price_C REAL)")
        cursor.execute(f"INSERT OR REPLACE INTO {table_name} (city, price_A, price_B, price_C) VALUES (?, ?, ?, ?)",
                       (city, *prices))
    else:
        cursor.execute(f"CREATE TABLE IF NOT EXISTS {table_name} (city TEXT PRIMARY KEY, price_A REAL, price_B REAL)")
        cursor.execute(f"INSERT OR REPLACE INTO {table_name} (city, price_A, price_B) VALUES (?, ?, ?)",
                       (city, *prices))

    conn.commit()
    conn.close()


# Сравнение цен по доставке и отправка сообщений об изменении цен
def compare_prices_and_update():
    conn = sqlite3.connect('bot_database.db')
    cursor = conn.cursor()

    cursor.execute("SELECT city, price_A, price_B, price_C FROM new_city_delivery")
    new_delivery_prices = cursor.fetchall()

    user_notifications = {}

    for city, price_A, price_B, price_C in new_delivery_prices:
        cursor.execute(
            "SELECT u.name, u.tg_link, uh.price_A, uh.price_B, uh.price_C, uh.user_id "
            "FROM user_city_history_delivery uh INNER JOIN users u ON uh.user_id = u.user_id WHERE uh.city=?",
            (city,))
        current_prices = cursor.fetchall()

        for name, tg_link, current_price_A, current_price_B, current_price_C, user_id in current_prices:
            if (current_price_A, current_price_B, current_price_C) != (price_A, price_B, price_C):
                cursor.execute(
                    "UPDATE user_city_history_delivery SET price_A=?, price_B=?, price_C=? WHERE city=? AND user_id=?",
                    (price_A, price_B, price_C, city, user_id))
                conn.commit()

                diff_A = round(price_A - current_price_A, 1)
                diff_B = round(price_B - current_price_B, 1)
                diff_C = round(price_C - current_price_C, 1)

                markup = types.InlineKeyboardMarkup()
                markup.row(
                    types.InlineKeyboardButton('Доставка', callback_data='доставка'),
                    types.InlineKeyboardButton('Контакты ☎️', callback_data='контакты')
                )

                notification_text_user = f"В городе {city}:\n\n"
                if diff_A != 0:
                    notification_text_user += f"🚛  Контейнером: {price_A} ₽ с НДС ({'+' if diff_A > 0 else ''}{diff_A:.1f} ₽)\n"
                else:
                    notification_text_user += f"🚛  Контейнером: {price_A} ₽ с НДС (0.0)\n"

                if diff_B != 0:
                    notification_text_user += f"🚚 На авто: {price_B} ₽ с НДС ({'+' if diff_B > 0 else ''}{diff_B:.1f} ₽)\n"
                else:
                    notification_text_user += f"🚚 На авто: {price_B} ₽ с НДС (0.0)\n"

                if diff_C != 0:
                    notification_text_user += f"🚂 Вагоном: {price_C} ₽ с НДС ({'+' if diff_C > 0 else ''}{diff_C:.1f} ₽)\n\n\n"
                else:
                    notification_text_user += f"🚂 Вагоном: {price_C} ₽ с НДС (0.0)\n\n\n"

                if user_id not in user_notifications:
                    user_notifications[user_id] = {"notification_text": "", "name": name}
                user_notifications[user_id]["notification_text"] += notification_text_user

    for user_id, user_data in user_notifications.items():
        if user_data["notification_text"]:
            bot.send_message(user_id, f'Цены на доставку были обновлены. \n\n{user_data["notification_text"]}',
                             parse_mode='HTML', reply_markup=markup)
            tg_username = bot.get_chat(user_id).username
            tg_link = f"https://t.me/{tg_username}"
            admin_notification_text = f"Пользователь {user_data['name']} " \
                                      f"(<a href='tg://user?id={user_id}'>{tg_link}</a>) получил " \
                                      f"сообщение об изменении цен:\n{user_data['notification_text']}\n\n"
            bot.send_message(admin_chat_id, admin_notification_text, parse_mode='HTML')

    conn.close()


# Сравнение цен по самовывозу и отправка сообщений об изменении цен
def compare_prices_and_update_pickup():
    conn = sqlite3.connect('bot_database.db')
    cursor = conn.cursor()

    cursor.execute("SELECT city, price_A, price_B FROM new_city_pickup")
    new_pickup_prices = cursor.fetchall()

    user_notifications = {}

    for city, price_A, price_B in new_pickup_prices:
        cursor.execute(
            "SELECT u.name, u.tg_link, uh.price_A, uh.price_B, uh.user_id "
            "FROM user_city_history_pickup uh INNER JOIN users u ON uh.user_id = u.user_id WHERE uh.city=?",
            (city,))
        current_pickup_prices = cursor.fetchall()

        for name, tg_link, current_price_A, current_price_B, user_id in current_pickup_prices:
            if (current_price_A, current_price_B) != (price_A, price_B):
                cursor.execute("UPDATE user_city_history_pickup SET price_A=?, price_B=? WHERE city=? AND user_id=?",
                               (price_A, price_B, city, user_id))
                conn.commit()

                diff_A = round(price_A - current_price_A, 1)
                diff_B = round(price_B - current_price_B, 1)

                markup = types.InlineKeyboardMarkup()
                markup.row(
                    types.InlineKeyboardButton('Самовывоз', callback_data='самовывоз'),
                    types.InlineKeyboardButton('Контакты ☎️', callback_data='контакты')
                )

                notification_text_user = f"В городе {city}:\n\n"
                if diff_A != 0:
                    notification_text_user += f"💰 Мешки 50кг: {price_A} ₽ с НДС ({'+' if diff_A > 0 else ''}{diff_A:.1f} ₽)\n"
                else:
                    notification_text_user += f"💰 Мешки 50кг: {price_A} ₽ с НДС (0.0)\n"

                if diff_B != 0:
                    notification_text_user += f"📦 Пачки 1кг, 5кг: {price_B} ₽ с НДС ({'+' if diff_B > 0 else ''}{diff_B:.1f} ₽)\n\n\n"
                else:
                    notification_text_user += f"📦 Пачки 1кг, 5кг: {price_B} ₽ с НДС (0.0)\n\n\n"

                if user_id not in user_notifications:
                    user_notifications[user_id] = {"notification_text": "", "name": name}
                user_notifications[user_id]["notification_text"] += notification_text_user

    for user_id, user_data in user_notifications.items():
        if user_data["notification_text"]:
            bot.send_message(user_id, f'Цены на самовывоз были обновлены. \n\n{user_data["notification_text"]}',
                             reply_markup=markup)
            tg_username = bot.get_chat(user_id).username
            tg_link = f"https://t.me/{tg_username}"
            admin_notification_text = f"Пользователь {user_data['name']} " \
                                      f"(<a href='tg://user?id={user_id}'>{tg_link}</a>) получил " \
                                      f"сообщение об изменении цен:\n{user_data['notification_text']}\n\n"
            bot.send_message(admin_chat_id, admin_notification_text, parse_mode='HTML')

    conn.close()


# Основные callback
@bot.callback_query_handler(func=lambda call: True)
def callback_handler(call):
    try:
        if call.data == 'самовывоз':
            send_cities(call.message.chat.id, call.message.message_id, call.data)
        elif call.data == 'доставка':
            send_cities(call.message.chat.id, call.message.message_id, call.data)
        elif call.data == 'контакты':
            send_cities(call.message.chat.id, call.message.message_id, call.data)
        elif call.data == 'back_to_cities':
            current_shipment_type = cities_data.get(call.message.chat.id, {}).get('shipment_type')
            if current_shipment_type:
                update_cities_message(call.message.chat.id, call.message.message_id, current_shipment_type)
            else:
                raise KeyError("Shipment type not found for the user")
        elif call.data == 'not_found':
            user_id = call.message.chat.id
            bot.send_message(user_id, "Пожалуйста, укажите ваш город:")
            bot.register_next_step_handler(call.message, forward_not_found_message)
        elif call.data == 'back_to_main_menu':
            main_menu(call.message)
        else:
            city_price_handler(call)
            add_city_to_history(call.message.chat.id, call.data,
                                cities_data.get(call.message.chat.id, {}).get('shipment_type'))
    except Exception as e:
        markup = types.InlineKeyboardMarkup(row_width=1)
        btn1 = types.InlineKeyboardButton('Самовывоз', callback_data='самовывоз')
        btn2 = types.InlineKeyboardButton('Доставка', callback_data='доставка')
        btn3 = types.InlineKeyboardButton('Контакты', callback_data='контакты')
        markup.add(btn1, btn2, btn3)

        bot.send_message(call.message.chat.id, "Выберете интересующий вас раздел",
                         reply_markup=markup)
        traceback.print_exc()


# Сообщения с кнопками городов или отделов после нажатия на кнопку "назад"
def update_cities_message(user_id, message_id, shipment_type):
    file_path = 'бот.xlsx'
    if shipment_type == 'самовывоз':
        sheet_name = 'самовывоз'
        message_text = "Выберите город для самовывоза:"
        add_not_found_button = True
        btn1 = None
    elif shipment_type == 'доставка':
        sheet_name = 'доставка'
        message_text = "Выберите город для доставки:"
        add_not_found_button = True
        btn1 = None
    elif shipment_type == 'контакты':
        sheet_name = 'контакты'
        message_text = "Выберите город или отдел:"
        add_not_found_button = False
        btn1 = None

    cities = read_cities_from_excel(file_path, sheet_name)
    num_rows = len(cities) // 2 + len(cities) % 2
    markup = types.InlineKeyboardMarkup(row_width=2)

    for i in range(num_rows):
        button1 = types.InlineKeyboardButton(cities[i * 2], callback_data=cities[i * 2])
        row = [button1]
        if i * 2 + 1 < len(cities):
            button2 = types.InlineKeyboardButton(cities[i * 2 + 1], callback_data=cities[i * 2 + 1])
            row.append(button2)
        markup.add(*row)

    if add_not_found_button:
        markup.add(types.InlineKeyboardButton("Тут нет моего города", callback_data="not_found"))

    if btn1:
        markup.add(btn1)

    markup.add(types.InlineKeyboardButton("Назад ⬅️", callback_data="back_to_main_menu"))
    cities_data[user_id] = {'message_id': message_id, 'shipment_type': shipment_type}
    bot.send_message(chat_id=user_id, text=message_text, reply_markup=markup)


# Клиент не нашел свой город
def forward_not_found_message(message):
    user_id = message.chat.id
    user_name = message.from_user.first_name
    tg_username = bot.get_chat(user_id).username
    tg_link = f"https://t.me/{tg_username}"
    user_mention = f"{user_name} (<a href='tg://user?id={user_id}'>{tg_link})</a>"

    markup = types.InlineKeyboardMarkup(row_width=1)
    btn1 = types.InlineKeyboardButton('Самовывоз', callback_data='самовывоз')
    btn2 = types.InlineKeyboardButton('Доставка', callback_data='доставка')
    btn3 = types.InlineKeyboardButton('Контакты', callback_data='контакты')
    markup.add(btn1, btn2, btn3)

    city = message.text
    bot.send_message(user_id, "По данному направлению просим ожидать расчета либо звонка от специалиста компании",
                     parse_mode='HTML', reply_markup=markup)

    data = cities_data.get(user_id)
    if data:
        shipment_type = data['shipment_type']
        if shipment_type == 'доставка':
            section = "разделе доставка"
        elif shipment_type == 'самовывоз':
            section = "разделе самовывоз"
        elif shipment_type == 'контакты':
            section = "разделе контакты"
        else:
            section = "неопределенном разделе"
    else:
        section = "неопределенном разделе"
    admin_message = f"Пользователь {user_mention} не нашел свой город в {section}. Его город: {city}"

    bot.send_message(admin_chat_id, admin_message, parse_mode='HTML')


# Получение городов из таблицы
def read_cities_from_excel(file_path, sheet_name):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    cities = []
    for cell in sheet['A']:
        if cell.value:
            cities.append(cell.value)
    return cities


# Сообщения с кнопками городов или отделов
def send_cities(user_id, message_id, shipment_type):
    file_path = 'бот.xlsx'
    if shipment_type == 'самовывоз':
        sheet_name = 'самовывоз'
        message_text = "Выберите город для самовывоза:"
        add_not_found_button = True
        btn1 = None
    elif shipment_type == 'доставка':
        sheet_name = 'доставка'
        message_text = "Выберите город для доставки:"
        add_not_found_button = True
        btn1 = None
    elif shipment_type == 'контакты':
        sheet_name = 'контакты'
        message_text = "Выберите город или отдел:"
        add_not_found_button = False
        btn1 = None

    cities = read_cities_from_excel(file_path, sheet_name)
    num_rows = len(cities) // 2 + len(cities) % 2
    markup = types.InlineKeyboardMarkup(row_width=2)

    for i in range(num_rows):
        button1 = types.InlineKeyboardButton(cities[i * 2], callback_data=cities[i * 2])
        row = [button1]
        if i * 2 + 1 < len(cities):
            button2 = types.InlineKeyboardButton(cities[i * 2 + 1], callback_data=cities[i * 2 + 1])
            row.append(button2)
        markup.add(*row)

    if add_not_found_button:
        markup.add(types.InlineKeyboardButton("Тут нет моего города", callback_data="not_found"))

    if btn1:
        markup.add(btn1)

    markup.add(types.InlineKeyboardButton("Назад ⬅️", callback_data="back_to_main_menu"))
    cities_data[user_id] = {'message_id': message_id, 'shipment_type': shipment_type}
    bot.send_message(user_id, message_text, reply_markup=markup)


# Получение городов для из листа контакты
def get_city_contacts(city):
    file_path = 'бот.xlsx'
    sheet_name = 'контакты'
    wb = load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(values_only=True):
        if row[0] == city:
            name = row[1] if row[1] is not None else "–"
            address = row[2]
            loading_contact = row[3] if row[3] is not None else "–"
            manager_contact = row[4] if row[4] is not None else "–"
            manager_email = row[5] if row[5] is not None else "–"
            telegram = row[6] if row[6] is not None else None
            whatsapp = row[7] if row[7] is not None else None

            if address:
                address_link = f"https://yandex.ru/maps/?text={address.replace(' ', '+')}"
                address_info = f"📌 Адрес: <a href='{address_link}'>{address}</a>"
            else:
                address_info = "📌 Адрес: –"

            contact_info = f"{city}:" \
                           f"\n\n{address_info}" \
                           f"\n👨‍💼 ФИО: {name}" \
                           f"\n☎️ Контакт на погрузке: <a href='tel:{loading_contact}'>{loading_contact}</a>" \
                           f"\n📲 Контакт менеджера: <a href='tel:{manager_contact}'>{manager_contact}</a>" \
                           f"\n📧 Почта менеджера: <a href='mailto:{manager_email}'>{manager_email}</a>"

            if whatsapp:
                contact_info += f"\n📱 Написать в <a href='{whatsapp}'>WhatsApp</a>"
            if telegram:
                contact_info += f"\n▶️ Написать в <a href='{telegram}'>Telegram</a>"

            return contact_info

    return "Информация о контактах для данного города отсутствует"


# Получение цен в городах
def get_city_price(city, shipment_type):
    if shipment_type == 'контакты':
        return get_city_contacts(city), None, None, None, None, None

    file_path = 'бот.xlsx'
    if shipment_type == 'самовывоз':
        sheet_name = 'самовывоз'
    elif shipment_type == 'доставка':
        sheet_name = 'доставка'

    wb = load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]

    for row in sheet.iter_rows(values_only=True):
        if row[0] == city:
            price_a = row[1] if row[1] is not None else '–'
            note_a = row[2] if len(row) > 2 and row[2] is not None else ''
            price_b = row[3] if row[3] is not None else '–'
            note_b = row[4] if len(row) > 4 and row[4] is not None else ''

            note_text_a = f" ({note_a})" if note_a else ""
            note_text_b = f" ({note_b})" if note_b else ""

            price_a_rounded = round(price_a, 1) if price_a != '–' else '–'
            price_b_rounded = round(price_b, 1) if price_b != '–' else '–'

            message = f"Текущее предложение на самовывоз с города {city}:\n\n" \
                      f"💰 Мешки 50кг: {price_a_rounded} ₽ с НДС{note_text_a}\n" \
                      f"📦 Пачки 1кг, 5кг: {price_b_rounded} ₽ с НДС{note_text_b}\n\n" \
                      f"Данное предложение носит информационный характер, для уточнения скидки и условий отгрузки, " \
                      f"а также для окончательного подтверждения просим связаться с закрепленным менеджером " \
                      f"либо по контактам компании. Возможна организация доставки."

            markup = types.InlineKeyboardMarkup(row_width=2)

            markup.add(
                types.InlineKeyboardButton('Доставка', callback_data='доставка'),
                types.InlineKeyboardButton('Контакты ☎️', callback_data='контакты'),
                types.InlineKeyboardButton("Самовывоз", callback_data="back_to_cities")
            )

            return message, markup, price_a_rounded, price_b_rounded, note_text_a, note_text_b

    return "Информация о стоимости для данного города отсутствует", None, None, None, None, None


# Отправка цен по доставке в городах
def send_delivery_price(user_id, message_id, city, container_price, auto_price, wagon_price, note_a, note_b, note_c, shipment_type):
    markup = types.InlineKeyboardMarkup()
    markup.row(
        types.InlineKeyboardButton('Самовывоз', callback_data='самовывоз'),
        types.InlineKeyboardButton('Контакты ☎️', callback_data='контакты'),
    )
    markup.row(
        types.InlineKeyboardButton("Доставка", callback_data="back_to_cities")
    )

    container_text = f"{round(float(container_price), 1)} ₽ с НДС" if container_price != "–" else "–"
    auto_text = f"{round(float(auto_price), 1)} ₽ с НДС" if auto_price != "–" else "–"
    wagon_text = f"{round(float(wagon_price), 1)} ₽ с НДС" if wagon_price != "–" else "–"

    note_text_a = f" ({note_a})" if note_a else ""
    note_text_b = f" ({note_b})" if note_b else ""
    note_text_c = f" ({note_c})" if note_c else ""

    message = f"Текущее предложение на {city}:\n\n" \
              f"🚛 Контейнером: {container_text}{note_text_a}\n" \
              f"🚚 На авто: {auto_text}{note_text_b}\n" \
              f"🚂 Вагоном: {wagon_text}{note_text_c}\n\n" \
              f"Данное предложение носит информационный характер, для уточнения скидки и условий отгрузки, а также " \
              f"для окончательного подтверждения просим связаться с закрепленным менеджером или по контактам компании."

    bot.send_message(chat_id=user_id, text=message, reply_markup=markup)

    first_name = bot.get_chat(user_id).first_name
    tg_username = bot.get_chat(user_id).username
    tg_link = f"https://t.me/{tg_username}"
    admin_message = f"Пользователь {first_name} (<a href='tg://user?id={user_id}'>{tg_link}</a>) с ID: <u>{user_id}</u> " \
                    f"интересовался ценами на доставку в городе {city}:\n\n" \
                    f"🚛 Контейнером: {container_text}{note_text_a}\n" \
                    f"🚚 На авто: {auto_text}{note_text_b}\n" \
                    f"🚂 Вагоном: {wagon_text}{note_text_c}"

    bot.send_message(admin_chat_id, admin_message, parse_mode='HTML')


# Отправка цен по самовывозу в городах
@bot.callback_query_handler(
    func=lambda call: call.data and call.data != 'next' and call.data != 'prev' and call.data != 'not_found')
def city_price_handler(call):
    user_id = call.message.chat.id
    data = cities_data.get(user_id)
    if data:
        city = call.data
        shipment_type = data['shipment_type']
        if shipment_type == 'доставка':
            file_path = 'бот.xlsx'
            sheet_name = 'доставка'
            wb = load_workbook(file_path, data_only=True)
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                if row[0] == city:
                    container_price = row[1] if row[1] is not None else "–"
                    note_a = row[2] if len(row) > 2 else ""
                    auto_price = row[3] if row[3] is not None else "–"
                    note_b = row[4] if len(row) > 4 else ""
                    wagon_price = row[5] if row[5] is not None else "–"
                    note_c = row[6] if len(row) > 6 else ""
                    send_delivery_price(user_id, call.message.message_id, city, container_price, auto_price,
                                        wagon_price, note_a, note_b, note_c, shipment_type)
                    return
            bot.send_message(user_id, "Информация о стоимости доставки для данного города отсутствует")
        elif shipment_type == 'контакты':
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton("Назад", callback_data="back_to_cities"))
            contact_info = get_city_contacts(city)
            bot.send_message(chat_id=user_id, text=contact_info, reply_markup=markup, parse_mode='HTML')
        else:
            price_info, markup, price_a, price_b, note_text_a, note_text_b = get_city_price(city, shipment_type)
            bot.send_message(chat_id=user_id, text=price_info, reply_markup=markup)

            first_name = bot.get_chat(user_id).first_name
            tg_username = bot.get_chat(user_id).username
            tg_link = f"https://t.me/{tg_username}"
            admin_message = f"Пользователь {first_name} (<a href='tg://user?id={user_id}'>{tg_link}</a>) с ID: <u>{user_id}</u> " \
                            f"интересовался ценами на самовывоз в городе {city}:\n\n"
            if shipment_type != 'доставка':
                admin_message += f"💰 Мешки 50кг: {price_a} ₽ с НДС{note_text_a}\n" \
                                 f"📦 Пачки 1кг, 5кг: {price_b} ₽ с НДС{note_text_b}\n\n"

            bot.send_message(admin_chat_id, admin_message, parse_mode='HTML')


# Проверка базы данных на пустые значения или числа с плавающей точкой
def check_and_remove_none_prices():
    conn = sqlite3.connect('bot_database.db')
    cursor = conn.cursor()

    tables_to_check = {
        'user_city_history_pickup': ['price_A', 'price_B'],
        'user_city_history_delivery': ['price_A', 'price_B', 'price_C'],
        'new_city_pickup': ['price_A', 'price_B'],
        'new_city_delivery': ['price_A', 'price_B', 'price_C']
    }

    for table, columns in tables_to_check.items():
        cursor.execute(f"SELECT rowid FROM {table} WHERE {' OR '.join(f'{col} IS NULL' for col in columns)}")
        rows_to_delete = cursor.fetchall()
        for row_id in rows_to_delete:
            cursor.execute(f"DELETE FROM {table} WHERE rowid=?", row_id)
            conn.commit()

        for column in columns:
            cursor.execute(f"SELECT rowid, {column} FROM {table}")
            rows = cursor.fetchall()
            for row in rows:
                if isinstance(row[1], float) and row[1] % 1 != 0:
                    updated_value = round(row[1], 1)
                    cursor.execute(f"UPDATE {table} SET {column}=? WHERE rowid=?", (updated_value, row[0]))
                    conn.commit()

    conn.close()


scheduler = BackgroundScheduler()

if __name__ == "__main__":
    scheduler.add_job(check_and_remove_none_prices, 'cron', minute='*', hour='*', misfire_grace_time=60)
    scheduler.start()

    try:
        bot.polling(none_stop=True)
    except Exception as e:
        print(f"Ошибка: {e}")
        time.sleep(15)
